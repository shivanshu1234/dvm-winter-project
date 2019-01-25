from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views import generic 
from .models import Post
from .forms import NewPostForm, NewCommentForm, EditPostForm, UserCreationFormWithEmail  
from .forms import EditUserForm, NewReportForm, PictureForm
from django.utils import timezone
from django.contrib.auth.decorators import login_required, user_passes_test
import datetime
from django.core.exceptions import PermissionDenied
from django.contrib.auth.models import User
import os
from django.core.mail import send_mail
from django.template.loader import get_template
from openpyxl import Workbook
from django.http import HttpResponse


class SignUp(generic.CreateView):
    form_class = UserCreationFormWithEmail
    success_url = reverse_lazy('login') 
    # reverse_lazy : because class attributes are loaded upon import.
    template_name = 'bookedApp/signup.html'


class Home(generic.ListView):
    template_name = 'bookedApp/home.html'
    context_object_name = 'latest_post_list'
    
    def get_queryset(self):
        return Post.objects.order_by('-date')


def PostDetails(request, post_id):    
    post = get_object_or_404(Post, id = post_id)
    current_user = request.user
    reported = False # Has the current user reported this post yet?
    for report in post.report_set.all():
        if report.reporting_user == current_user:
            reported = True
    context = {
        'post':post,
        'reported':reported
        }
    
    return render(request, 'bookedApp/PostDetails.html', context = context)


@login_required
def NewPost(request):
    user = request.user
    if request.method == "POST":
        form = NewPostForm(request.POST)
        if form.is_valid():
            post = form.save(commit=False) 
            # commit=False : Create the instance but don't save it to the database.
            post.user = user
            post.date = timezone.now()
            post.save()

            subject = "New post on Booked! By " + user.username
            from_email = 'shivanshuayachi@gmail.com'
            
            # A list of IDs of all followers who have Email enabled for this user.
            EmailFollowersIDs = [dict['user'] for dict in user.EmailFollowers.all().values('user')]
            
            # A list of all their emails.
            to_email = [User.objects.filter(id=id).first().email for id in EmailFollowersIDs]
    
            context = {
                'user':post.user,
                'heading':post.heading
            }

            message = get_template('bookedApp/new_post_email.txt').render(context)
            send_mail(subject, message, from_email, to_email, fail_silently = False)

            return redirect('bookedApp:home')
    else:
        form = NewPostForm()
    
    return render(request, template_name = 'bookedApp/NewPost.html', context={'form':form})


@login_required
def NewComment(request, post_id):
    post = get_object_or_404(Post, id = post_id)
    if request.method == "POST":
        form = NewCommentForm(request.POST)
        if form.is_valid():
            comment = form.save(commit = False)
            comment.post = post
            comment.user = request.user
            comment.save()
            return redirect('bookedApp:PostDetails', post_id = post_id)

    else:
        form  = NewCommentForm()
    
    context={
        'form':form, 
        'post':post
        }
    
    return render(request, template_name = 'bookedApp/NewComment.html', context = context)


@login_required
def MyPosts(request):
    user = request.user
    # Posts in reverse chronological order.
    posts_by_user = Post.objects.filter(user = user).order_by('-date')
    
    context = {
        'user':user, 
        'posts_by_user':posts_by_user
        }
    
    return render(request, template_name = "bookedApp/MyPosts.html", context = context)


@login_required
def EditPost(request, post_id):
    post = get_object_or_404(Post, id = post_id)
    
    if timezone.now() - post.date < datetime.timedelta(minutes = 5):
        # Posts are editable only till 5 minutes after posting them.
        post_is_editable = True  
        
        if post.user == request.user:
            if request.method == "POST":
                form = EditPostForm(request.POST,)
                if form.is_valid():
                    post = form.save(commit = False)
                    post.id = post_id
                    post.user = request.user
                    post.date = timezone.now()
                    
                    # Append "edited" to the heading if it's not already there.
                    if not post.heading.endswith('(edited)'):  
                        post.heading = post.heading + "(edited)"
                    else: 
                        post.heading = post.heading
                    
                    post.save()
                    return redirect('bookedApp:MyPosts')
            else:
                form = EditPostForm(initial={'heading':post.heading, 'text':post.text})
        
        context = {
            'form':form, 
            'post_is_editable':post_is_editable
            }

        return render(request, template_name = 'bookedApp/EditPost.html', context=context)
                
    else: 
        post_is_editable = False # If more than 5 minutes have elapsed since posting, disable editing.
        context = {'post_is_editable':post_is_editable}               
        return render(request, template_name = 'bookedApp/EditPost.html', context=context)


def ConfirmDeletePost(request, post_id):
    post = get_object_or_404(Post, id = post_id)
    if post.user == request.user:
        
        context = {
            'heading':post.heading, 
            'id':post.id
            }
        
        return render(request, template_name = 'bookedApp/ConfirmDeletePost.html', context=context)
    else:
        raise PermissionDenied


def DeletePost(request, post_id):
    post = get_object_or_404(Post, id = post_id)
    if post.user == request.user:
        post.delete()
        return redirect('bookedApp:PostDeleted')
    else: 
        raise PermissionDenied


def PostDeleted(request):
    return render(request, template_name = 'bookedApp/PostDeleted.html', context = {})


@login_required
def MyProfile(request):
    user = request.user
    return render(request, template_name = 'bookedApp/MyProfile.html', context = {'user':user})


@login_required
def EditProfile(request):
    user = request.user
    if request.method == "POST":
        form = EditUserForm(request.POST, instance=user)
        
        if form.is_valid():
            user.email = request.POST['email']
            user.username = form.clean_username()
            user.save()
            return redirect('bookedApp:UserEdited')

    else:
        form = EditUserForm(initial = {'username':user.username, 'email':user.email})
        # 'initial = ' pre-populates the form with the current data of the user.

    context = {'form':form}
    return render(request, template_name = 'bookedApp/EditProfile.html', context = context)


def UserEdited(request):
    return render(request, template_name = 'bookedApp/UserEdited.html', context = {})


def ViewProfile(request, user_id):
    current_user = request.user
    other_user = get_object_or_404(User, id = user_id)
    posts = Post.objects.filter(user = other_user).order_by('-date')

    # Is the current user viewing his own profile?
    if current_user == other_user:
        own_profile = True
    else:
        own_profile = False
    
    # Does the current user follow this user?
    if other_user in current_user.userprofile.follows.all():
        user_follows_other_user = True
    else:
        user_follows_other_user = False

    # Has the current user enabled the Email notifications for this user.
    if other_user in current_user.userprofile.EmailFollows.all():
        emails_enabled = True
    else:
        emails_enabled = False
    
    context = {
        'other_user': other_user, 
        'user_follows_other_user':user_follows_other_user,
        'posts':posts,
        'own_profile':own_profile,
        'emails_enabled':emails_enabled,
        }

    return render(request, template_name = 'bookedApp/ViewProfile.html', context = context)


@login_required
def Follow(request, user_id):
    user_to_be_followed = get_object_or_404(User, id = user_id)
    current_user = request.user
    # Add current user to this user's followers list.
    current_user.userprofile.follows.add(user_to_be_followed)
    return redirect('ViewProfile/')


@login_required
def UnFollow(request, user_id):
    user_to_be_unfollowed = get_object_or_404(User, id = user_id)
    current_user = request.user

    # Remove user from current user's followers list.
    if user_to_be_unfollowed in current_user.userprofile.follows.all():
        current_user.userprofile.follows.remove(user_to_be_unfollowed)
    
    # Also disable email notifications, if previously enabled.
    if user_to_be_unfollowed in current_user.userprofile.EmailFollows.all():
        current_user.userprofile.EmailFollows.remove(user_to_be_unfollowed)
    
    return redirect('ViewProfile/')


@login_required
def Followers(request):
    user = request.user
    # All users following the current user.
    followers = user.followers.all()
    context = {'followers':followers}
    return render(request, template_name = 'bookedApp/Followers.html', context = context)


@login_required
def Following(request):
    user = request.user
    # All users followed by the current user.
    following = user.userprofile.follows.all()
    context = {'following':following}
    return render(request, template_name = 'bookedApp/Following.html', context = context)


@login_required
def MyFeed(request):
    user = request.user
    # Posts by all the users that the current user follows.
    posts = Post.objects.filter(user__userprofile__follows__id = user.id)
    context = {'posts':posts}
    return render(request, template_name = 'bookedApp/MyFeed.html', context = context)


@login_required
def ReportPost(request, user_id, post_id):
    post = get_object_or_404(Post, id = post_id)
    if request.method == "POST":
        form = NewReportForm(request.POST)
        if form.is_valid():
            report = form.save(commit = False)
            report.post = post
            report.reported_user = get_object_or_404(User, id = user_id)
            report.reporting_user = request.user
            report.save()

            if post.report_set.all().count() > 2: # If post has more than 2 reports, remove it.
                post.delete()
                return redirect("/bookedApp/")

            return redirect("/bookedApp/"+str(post_id)+"/")  # "/bookedApp/<post_id>/"
    else:
        form = NewReportForm()
    
    context = {
        'form':form,
        'post':post
        }
  
    return render(request, template_name = 'bookedApp/ReportPost.html', context = context)


@login_required
def ChangeProfilePicture(request):
    user = request.user
    old_pic = user.userprofile.picture
    if request.method == "POST":
        form = PictureForm(request.POST, request.FILES, instance = user.userprofile)
        if form.is_valid():
            profile = form.save(commit = False)
            
            try: # If a user clears his picture, reset it to the default.
                profile.picture.url
            except ValueError:
                profile.picture = 'default.jpeg'
            
            profile.save()
            
            if not old_pic in ['default.jpeg', profile.picture]:
                # If the old pic is not the same as the new one or the default,
                # delete the old pic. 
                os.remove(old_pic.path)
                
            return redirect('bookedApp:MyProfile')

    else:
        form = PictureForm(instance = user.userprofile)
       
        context = {'form':form}

    return render(request, template_name = 'bookedApp/ChangeProfilePicture.html', context = context)


@login_required
def EmailFollow(request, user_id):
    user_to_be_followed = get_object_or_404(User, id = user_id)
    current_user = request.user
    
    # Enable the emails.
    current_user.userprofile.EmailFollows.add(user_to_be_followed)
    return redirect('ViewProfile/')


@login_required
def EmailUnFollow(request, user_id):
    user_to_be_unfollowed = get_object_or_404(User, id = user_id)
    current_user = request.user
    
    # If the user has enabled emails, disable them.
    if user_to_be_unfollowed in current_user.userprofile.EmailFollows.all():
        current_user.userprofile.EmailFollows.remove(user_to_be_unfollowed)
    
    return redirect('ViewProfile/')

@user_passes_test(lambda u: u.is_superuser) # Check if user is a superuser.
def GenerateExcelSheet(request, SelectedIDs):
    SelectedIDs = SelectedIDs.split('-') # Convert the string obtained from url into a list.
    workbook = Workbook()
    worksheet = workbook.active
    
    # Create the headings for the columns.
    worksheet['A1'] = 'username'
    worksheet['B1'] = 'id'
    worksheet['C1'] = 'email'

    # Enter data in the worksheet, iterating through user's IDs.
    for id in SelectedIDs:
        user = get_object_or_404(User, id = id) 
        users_index_in_SelectedIDs = SelectedIDs.index(id)
        
        worksheet.cell(row = users_index_in_SelectedIDs + 3, column = 1, value = user.username)
        worksheet.cell(row = users_index_in_SelectedIDs + 3, column = 2, value = user.id)
        worksheet.cell(row = users_index_in_SelectedIDs + 3, column = 3, value = user.email)

    response = HttpResponse(content_type = 'application/ms-excel')
    workbook.save(response)
    
    response['Content-Disposition'] = 'attachment; filename = user_data.xlsx'
    return response